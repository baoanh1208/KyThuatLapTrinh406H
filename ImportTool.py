from openpyxl.reader.excel import load_workbook

from K24406H.DoAn.models.Employee import Employee
from K24406H.DoAn.models.Material import Material


def import_material_excel(self, filename):
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    is_header = True
    materials = []
    for row in ws.values:
        if is_header == True:
            is_header = False
            continue
        name = row[0]
        type = row[1]
        supplier = row[2]
        import_price = row[3]
        import_qty=row[4]
        sell_price=row[5]
        sold_qty=row[6]
        use_day=row[7]
        m = Material(name,type,supplier,import_price,import_qty,sell_price,sold_qty,use_day)
        materials.append(m)
    wb.close()
    return materials


def import_employee_excel(self, filename):
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    is_header = True
    employees = []
    for row in ws.values:
        if is_header == True:
            is_header = False
            continue
        EmployeeId = row[0]
        EmployeeName = row[1]
        UserName = row[2]
        Password = row[3]
        emp = Employee(EmployeeId, EmployeeName,UserName,Password)
        employees.append(emp)
    wb.close()
    return employees